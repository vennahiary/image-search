from flask import Flask, request, send_file, render_template_string
import requests
import os
import re
from io import BytesIO
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage
import tempfile
import zipfile

app = Flask(__name__)

CLIENT_ID = os.environ.get("NAVER_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("NAVER_CLIENT_SECRET", "")

HTML = '''
<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>이미지 자동 검색</title>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Malgun Gothic', sans-serif; background: #f5f5f5; display: flex; justify-content: center; align-items: center; min-height: 100vh; }
        .card { background: white; border-radius: 12px; padding: 40px; width: 480px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }
        h1 { font-size: 22px; margin-bottom: 8px; color: #222; }
        p.desc { font-size: 13px; color: #888; margin-bottom: 30px; }
        label { font-size: 13px; font-weight: bold; color: #444; display: block; margin-bottom: 6px; }
        input[type=file] { width: 100%; padding: 10px; border: 2px dashed #ddd; border-radius: 8px; font-size: 13px; cursor: pointer; margin-bottom: 20px; }
        input[type=file]:hover { border-color: #4f8ef7; }
        button { width: 100%; padding: 14px; background: #4f8ef7; color: white; border: none; border-radius: 8px; font-size: 15px; font-weight: bold; cursor: pointer; }
        button:hover { background: #3a7be0; }
        .progress { display: none; margin-top: 20px; }
        .bar-wrap { background: #eee; border-radius: 20px; height: 10px; margin-top: 8px; }
        .bar { background: #4f8ef7; height: 10px; border-radius: 20px; width: 0%; transition: width 0.3s; }
        .status { font-size: 13px; color: #666; margin-top: 8px; }
        .result { display: none; margin-top: 20px; padding: 16px; background: #f0f7ff; border-radius: 8px; text-align: center; }
        .result a { color: #4f8ef7; font-weight: bold; text-decoration: none; font-size: 15px; }
    </style>
</head>
<body>
<div class="card">
    <h1>🔍 이미지 자동 검색</h1>
    <p class="desc">이카운트 엑셀 파일을 올리면 모델명으로 이미지를 자동으로 찾아 넣어드려요.</p>
    <form id="uploadForm">
        <label>엑셀 파일 선택</label>
        <input type="file" id="fileInput" accept=".xlsx,.xls" required>
        <button type="submit">🚀 이미지 검색 시작</button>
    </form>
    <div class="progress" id="progress">
        <div class="status" id="statusText">처리 중...</div>
        <div class="bar-wrap"><div class="bar" id="bar"></div></div>
    </div>
    <div class="result" id="result">
        ✅ 완료! <a id="downloadLink" href="#">결과 파일 다운로드</a>
    </div>
</div>
<script>
document.getElementById('uploadForm').onsubmit = async function(e) {
    e.preventDefault();
    const file = document.getElementById('fileInput').files[0];
    if (!file) return;

    document.getElementById('progress').style.display = 'block';
    document.getElementById('result').style.display = 'none';
    document.getElementById('bar').style.width = '10%';
    document.getElementById('statusText').textContent = '파일 업로드 중...';

    const formData = new FormData();
    formData.append('file', file);

    document.getElementById('bar').style.width = '30%';
    document.getElementById('statusText').textContent = '이미지 검색 중... (잠시 기다려주세요)';

    try {
        const res = await fetch('/process', { method: 'POST', body: formData });
        document.getElementById('bar').style.width = '100%';

        if (res.ok) {
            const blob = await res.blob();
            const url = URL.createObjectURL(blob);
            document.getElementById('downloadLink').href = url;
            document.getElementById('downloadLink').download = '결과_' + file.name;
            document.getElementById('result').style.display = 'block';
            document.getElementById('statusText').textContent = '완료!';
        } else {
            const err = await res.text();
            document.getElementById('statusText').textContent = '오류: ' + err;
        }
    } catch(e) {
        document.getElementById('statusText').textContent = '오류가 발생했습니다.';
    }
};
</script>
</body>
</html>
'''

def extract_model_code(raw_text):
    if not raw_text or str(raw_text).strip() == "nan":
        return None
    text = str(raw_text).strip()
    text = re.sub(r'\([^)]*\)', '', text)
    text = re.split(r'[가-힣]', text)[0].strip()
    return text if text else None

def search_image(model_code):
    url = "https://openapi.naver.com/v1/search/shop.json"
    headers = {
        "X-Naver-Client-Id": CLIENT_ID,
        "X-Naver-Client-Secret": CLIENT_SECRET,
    }
    params = {"query": model_code, "display": 5}
    try:
        res = requests.get(url, headers=headers, params=params, timeout=5)
        data = res.json()
        items = data.get("items", [])
        if items:
            return items[0].get("image")
    except:
        pass
    return None

def download_image(img_url):
    try:
        res = requests.get(img_url, timeout=10)
        img = Image.open(BytesIO(res.content)).convert("RGB")
        w, h = img.size
        min_side = min(w, h)
        left = (w - min_side) // 2
        top = (h - min_side) // 2
        img = img.crop((left, top, left + min_side, top + min_side))
        img = img.resize((300, 300), Image.LANCZOS)
        buf = BytesIO()
        img.save(buf, "JPEG", quality=90)
        buf.seek(0)
        return buf
    except:
        return None

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/process', methods=['POST'])
def process():
    RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    file = request.files['file']
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # 행높이/열너비 설정
    ROW_HEIGHT = 45
    ws.column_dimensions["G"].width = 7

    tmp_dir = tempfile.mkdtemp()

    for row_num in range(3, ws.max_row + 1):
        raw_value = ws.cell(row=row_num, column=8).value
        if not raw_value:
            continue
        model_code = extract_model_code(raw_value)
        if not model_code:
            continue

        img_url = search_image(model_code)
        ws.row_dimensions[row_num].height = ROW_HEIGHT

        if not img_url:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = RED_FILL
            continue

        img_buf = download_image(img_url)
        if not img_buf:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=col).fill = RED_FILL
            continue

        # 이미지 저장 후 삽입
        img_path = os.path.join(tmp_dir, f"{row_num}.jpg")
        with open(img_path, 'wb') as f:
            f.write(img_buf.read())

        xl_img = XLImage(img_path)
        # 행높이 55pt = 73px, 열너비 8 = 60px
        xl_img.width = 50
        xl_img.height = 50
        ws.add_image(xl_img, f"G{row_num}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='결과.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True)
